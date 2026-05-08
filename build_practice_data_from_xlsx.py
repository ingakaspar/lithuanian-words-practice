#!/usr/bin/env python3
"""Build nouns + verbs JSON for the app from words.xlsx."""
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
LT_NLP = ROOT / "Lithuanian-nlp-tools"
sys.path.insert(0, str(LT_NLP))
os.chdir(LT_NLP)
from conjugator import conjugate  # noqa: E402
from decliner import decline_noun  # noqa: E402

XLSX_PATH = ROOT / "words.xlsx"
TEST_SHEET = "📝 Vocabulary Test"

# Strip lexical stress marks only; keep length marks so ū remains distinct from u.
_STRESS_ORDS = frozenset(
    [0x0300, 0x0301, 0x0302, 0x0303] + [0x0306, 0x030B, 0x030F, 0x0311, 0x0341, 0x0342]
)
_MEANINGS_5 = [
    "first person singular",
    "second person singular",
    "third person",
    "first person plural",
    "second person plural",
]


def strip_stress_marks(s: str) -> str:
    if not s:
        return s
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if ord(ch) not in _STRESS_ORDS)
    return unicodedata.normalize("NFC", s)


def strip_lt(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = "".join(ch for ch in s if re.match(r"[A-Za-ząčęėįšųūžĄČĘĖĮŠŲŪŽ-]", ch))
    return s.lower()


def scrub_stress(obj):
    if isinstance(obj, str):
        return strip_stress_marks(obj)
    if isinstance(obj, dict):
        return {k: scrub_stress(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [scrub_stress(x) for x in obj]
    return obj


def clean_word_cell(cell: object) -> str:
    s = str(cell or "").strip()
    if not s:
        return ""
    # Keep only the primary headword for drills.
    s = re.split(r"[,;/]", s)[0].strip()
    s = re.sub(r"\([^)]*\)", "", s).strip()
    return strip_stress_marks(s)


def parse_tier_group(cell: object) -> int:
    m = re.search(r"(\d+)", str(cell or ""))
    if not m:
        return 1
    return int(m.group(1))


def tense_six(d: object) -> list:
    if not isinstance(d, dict):
        return ["", "", "", "", "", ""]
    vals = [(d.get(k) or "").strip() for k in _MEANINGS_5]
    third = vals[2]
    return [vals[0], vals[1], third, vals[3], vals[4], third]


def conditional_six(d: object) -> list:
    return tense_six(d)


def imperative_six(d: object) -> list:
    if not isinstance(d, dict):
        return ["", "", "", "", "", ""]
    return [
        "",
        (d.get("second person singular") or "").strip(),
        "",
        (d.get("first person plural") or "").strip(),
        (d.get("second person plural") or "").strip(),
        "",
    ]


def decl_ok(d: object) -> bool:
    if not d or not isinstance(d, dict):
        return False
    sg = d.get("Singular")
    return bool(sg and isinstance(sg, dict) and sg.get("Nominative"))


def load_words_from_xlsx() -> list:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if TEST_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet {TEST_SHEET!r} in {XLSX_PATH}")
    ws = wb[TEST_SHEET]
    out = []
    # B:# C:word E:tier G:translation
    for row in range(2, ws.max_row + 1):
        lt_raw = clean_word_cell(ws.cell(row, 3).value)
        if not lt_raw:
            continue
        lt_ascii = strip_lt(lt_raw)
        if len(lt_ascii) < 2:
            continue
        tier_group = parse_tier_group(ws.cell(row, 5).value)
        tr = str(ws.cell(row, 7).value or "").strip()
        out.append(
            {
                "lt": lt_raw,
                "lt_ascii": lt_ascii,
                "ru": tr,
                "group": tier_group,
            }
        )
    return out


def build_verbs(rows: list) -> tuple[list, set]:
    merged = {}
    for r in rows:
        key = r["lt_ascii"]
        if key not in merged:
            merged[key] = {"lt": r["lt"], "lt_ascii": key, "ru": r["ru"], "groups": {r["group"]}}
        else:
            merged[key]["groups"].add(r["group"])
            if not merged[key]["ru"] and r["ru"]:
                merged[key]["ru"] = r["ru"]

    out = []
    used_groups = set()
    for v in merged.values():
        lt = v["lt"]
        if not (lt.endswith("ti") or lt.endswith("tis")):
            continue
        try:
            c = conjugate(lt)
        except (KeyError, IndexError, TypeError, ValueError):
            continue
        present = c.get("present") or {}
        if not (present.get("third person") or "").strip():
            continue
        groups = sorted(v["groups"])
        used_groups.update(groups)
        out.append(
            {
                "lt": lt,
                "lt_ascii": v["lt_ascii"],
                "ru": v["ru"],
                "present": tense_six(c.get("present")),
                "past": tense_six(c.get("past")),
                "future": tense_six(c.get("future")),
                "usedTo": tense_six(c.get("past iterative")),
                "participlePastMasc": "",
                "conditional": conditional_six(c.get("conditional")),
                "imperative": imperative_six(c.get("imperative")),
                "groups": groups,
            }
        )
    out.sort(key=lambda x: x["lt_ascii"])
    return out, used_groups


def build_nouns(rows: list) -> tuple[list, set]:
    merged = {}
    for r in rows:
        key = r["lt_ascii"]
        if key not in merged:
            merged[key] = {"lt": r["lt"], "lt_ascii": key, "ru": r["ru"], "groups": {r["group"]}}
        else:
            merged[key]["groups"].add(r["group"])
            if not merged[key]["ru"] and r["ru"]:
                merged[key]["ru"] = r["ru"]

    out = []
    used_groups = set()
    for n in merged.values():
        try:
            decl = decline_noun(n["lt"])
        except (KeyError, IndexError, TypeError):
            continue
        if not decl_ok(decl):
            continue
        sg = decl["Singular"]
        if not all(k in sg for k in ("Nominative", "Genitive", "Vocative")):
            continue
        nom = strip_stress_marks(str(sg["Nominative"]))
        groups = sorted(n["groups"])
        used_groups.update(groups)
        out.append(
            {
                "lt": nom,
                "lt_ascii": strip_lt(nom),
                "ru": n["ru"],
                "groups": groups,
                "decl": scrub_stress(decl),
            }
        )
    out.sort(key=lambda x: x["lt_ascii"])
    return out, used_groups


def main() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing input file: {XLSX_PATH}")

    rows = load_words_from_xlsx()
    verbs, verb_groups = build_verbs(rows)
    nouns, noun_groups = build_nouns(rows)
    groups = sorted(verb_groups | noun_groups)

    verbs_path = ROOT / "conjugations_group0_4.json"
    nouns_path = ROOT / "nouns_group0_4.json"
    groups_path = ROOT / "pdf_group_ids.json"
    verbs_path.write_text(json.dumps(verbs, ensure_ascii=False, indent=2), encoding="utf-8")
    nouns_path.write_text(json.dumps(nouns, ensure_ascii=False, indent=2), encoding="utf-8")
    groups_path.write_text(json.dumps(groups, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(
        f"xlsx rows: {len(rows)} -> verbs: {len(verbs)} ({verbs_path}), "
        f"nouns: {len(nouns)} ({nouns_path}), groups: {groups} ({groups_path})"
    )


if __name__ == "__main__":
    main()
